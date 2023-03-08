import mechanize
import re
import requests
import sys
from bs4 import BeautifulSoup
import openpyxl
from lxml import html
import os
import ssl
import tkinter as tk


# Disable SSL certificate verification globally
os.environ['PYTHONHTTPSVERIFY'] = '0'

class LoginForm:
    def __init__(self, master):
        self.master = master
        master.title("Login Form")

        self.label_username = tk.Label(master, text="Username")
        self.label_username.pack()

        self.entry_username = tk.Entry(master)
        self.entry_username.pack()

        self.label_password = tk.Label(master, text="Password")
        self.label_password.pack()

        self.entry_password = tk.Entry(master, show="*")
        self.entry_password.pack()

        self.submit_button = tk.Button(master, text="Submit", command=self.submit)
        self.submit_button.pack()
        
        # Define instance variables to store username and password
        self.username = None
        self.password = None

    def submit(self):
        self.username = self.entry_username.get()
        self.password = self.entry_password.get()

        #print(f"Username: {username}")
        #print(f"Password: {password}")
        self.master.destroy()


def check_login(username:str, password:str) -> bool:
    ssl._create_default_https_context = ssl._create_unverified_context
    br = mechanize.Browser()

    # ignore robots.txt rules
    br.set_handle_robots(False)

    # navigate to the login page
    br.open('https://www.gpro.net/gb/Login.asp')

    # select the login form and fill in the fields
    br.select_form(nr=0)
    br.form['textLogin'] = username
    br.form['textPassword'] = password

    # submit the login form
    response = br.submit()
    
    response = list(br.links(url_regex=re.compile("DriverProfile")))
    
    # check the response for an error message indicating invalid credentials
    if len(response) == 1:
        print('Login successful.')
        return True
    # handle the error, e.g. retry with different credentials or exit the program
    else:
        print('Invalid username or password.')
        return False

def browser_open(username:str, password:str):
    ssl._create_default_https_context = ssl._create_unverified_context
    br = mechanize.Browser()

    # ignore robots.txt rules
    br.set_handle_robots(False)

    # navigate to the login page
    br.open('https://www.gpro.net/gb/Login.asp')

    # select the login form and fill in the fields
    br.select_form(nr=0)
    br.form['textLogin'] = username
    br.form['textPassword'] = password

    # submit the login form
    br.submit()
    
    return br

def extract_row_data(tree, first_arg:str, second_arg:str, third_arg:str, row_num:str):
    return tree.xpath(f"//{first_arg}[contains(@{second_arg}, '{third_arg}')]/tr[{row_num}]/td[1]/text()")[0]
    
def open_workbook_worksheet(workbook_path:str , worksheet_name:str):
    
    workbook = openpyxl.load_workbook(workbook_path)
    worksheet = workbook[worksheet_name]
    return workbook, worksheet  
    
def save_workbook(workbook, workbook_path):
    
    workbook.save(workbook_path)
    
def write_to_excel(data, gp_list, actual_gp:str , worksheet, start_row, end_row):
    
    
    gp_index = gp_list.index(actual_gp)
    col_num = gp_index + 1
    
    for item in data.items():
        worksheet.cell(row=start_row, column=col_num).value = item[1]
        start_row = start_row + 1
        if start_row == end_row:
            break

def fill_season_calendar():
    # send a GET request to the calendar page
    response = requests.get("https://www.gpro.net/gb/Calendar.asp")

    # create a BeautifulSoup object
    soup = BeautifulSoup(response.content, "html.parser")

    gp_list = []
    for row in soup.find_all('tr'):
        cols = row.find_all('td')
        if len(cols) >= 3:
            gp_name = cols[2].text.strip()
            gp_name = gp_name.split(' GP')[0]
            gp_list.append(gp_name)
            
    # Build a dictionary that maps GP names to their corresponding numbers
    gp_dict = {}
    for i, gp_name in enumerate(gp_list, start=1):
        gp_dict[gp_name] = str(i)

    # Find the element with class 'yellow'
    yellow_elem = soup.find(class_='yellow')

    # Get the text content of the element, strip leading/trailing white space
    yellow_text = yellow_elem.get_text().strip()

    # Remove any '0' characters and '.' characters from the text
    yellow_text = yellow_text.replace('0', '').replace('.', '')

    # Look up the GP name in the dictionary based on the modified text
    gp_name = next((name for name, num in gp_dict.items() if num == yellow_text), None)

    # Print the GP name if it was found, or a message if it wasn't found
    if gp_name:
        return gp_list, gp_name
    else:
        print(f"No GP name found for '{yellow_text}'")
        sys.exit()
    

def fill_driver_profile(br, gp_list , actual_gp):
    
    # navigate to the login page
    br.open('https://www.gpro.net/gb/DriverContract.asp')
    
    tree = html.fromstring(br.response().get_data())
    
    driver_data = {
        'overall': tree.xpath("normalize-space(//tr[contains(@data-step, '2')]//td/text())"),
        'concentration': extract_row_data(tree, 'table', 'class', 'squashed leftalign', '3'),
        'talent': extract_row_data(tree, 'table', 'class', 'squashed leftalign', '4'),
        'aggressiveness': extract_row_data(tree, 'table', 'class', 'squashed leftalign', '5'),
        'experience': extract_row_data(tree, 'table', 'class', 'squashed leftalign', '6'),
        'technical_insight': extract_row_data(tree, 'table', 'class', 'squashed leftalign', '7'),
        'stamina': extract_row_data(tree, 'table', 'class', 'squashed leftalign', '8'),
        'charisma': extract_row_data(tree, 'table', 'class', 'squashed leftalign', '9'),
        'motivation': extract_row_data(tree, 'table', 'class', 'squashed leftalign', '10'),
        'reputation': extract_row_data(tree, 'table', 'class', 'squashed leftalign', '12'),
        'weight': extract_row_data(tree, 'table', 'class', 'squashed leftalign', '14'),
        'age': extract_row_data(tree, 'table', 'class', 'squashed leftalign', '15')
    }
    
    workbook, worksheet = open_workbook_worksheet('me_s80.xlsx', '2 Stuff Data')
    
    write_to_excel(driver_data, gp_list, actual_gp, worksheet, 5, 17)
    # Load the workbook
    save_workbook(workbook, 'me_s80.xlsx')


def main():
    
    root = tk.Tk()
    login_form = LoginForm(root)
    root.mainloop()
    
    #check if the login is good
    if not check_login(login_form.username, login_form.password):
        sys.exit()
    
    browser = browser_open(login_form.username, login_form.password)
    
    gp_list, actual_gp = fill_season_calendar()
    
    fill_driver_profile(browser, gp_list , actual_gp)
    
    ### Until here is working with the new refactor

    # send a GET request to the calendar page
    response = requests.get("https://www.gpro.net/gb/Calendar.asp")

    # create a BeautifulSoup object
    soup = BeautifulSoup(response.content, "html.parser")

    gp_list = []
    for row in soup.find_all('tr'):
        cols = row.find_all('td')
        if len(cols) >= 3:
            gp_name = cols[2].text.strip()
            gp_name = gp_name.split(' GP')[0]
            gp_list.append(gp_name)
            

    print(gp_list)

    # Load the workbook
    workbook = openpyxl.load_workbook('me_s80.xlsx')

    worksheet = workbook['1 Season']
    for i in range(len(gp_list)):
        worksheet.cell(row=i+6, column=2).value = gp_list[i]

    # Save the changes to the workbook
    workbook.save('me_s80.xlsx')


    # Load the workbook
    workbook = openpyxl.load_workbook('me_s80.xlsx')

    # Select the worksheet
    worksheet = workbook['2 Stuff Data']

    #This needed to be changed, cannot be hardcoded
    current_gp = "Sakhir"

    gp_index = gp_list.index(current_gp)
    col_num = gp_index + 1  # add 2 to account for the header row

    print(col_num)

    # Write the data to the appropriate cell
    worksheet.cell(row=5, column=col_num).value = driverOverall
    worksheet.cell(row=6, column=col_num).value = driverConcentration
    worksheet.cell(row=7, column=col_num).value = driverTalent
    worksheet.cell(row=8, column=col_num).value = driverAggressiveness
    worksheet.cell(row=9, column=col_num).value = driverExperience
    worksheet.cell(row=10, column=col_num).value = driverTechnicalInsight
    worksheet.cell(row=11, column=col_num).value = driverStamina
    worksheet.cell(row=12, column=col_num).value = driverCharisma
    worksheet.cell(row=13, column=col_num).value = driverMotivation
    worksheet.cell(row=14, column=col_num).value = driverReputation
    worksheet.cell(row=15, column=col_num).value = driverWeight
    worksheet.cell(row=16, column=col_num).value = driverAge

    # Save the changes to the workbook
    workbook.save('me_s80.xlsx')

    # Staff and facilities is the same way that the driver with the steps

    response = br.open('https://www.gpro.net/gb/StaffAndFacilities.asp')

    #print(response)
    tree = html.fromstring(br.response().get_data())

    #Staff stuff
    StaffOverall = int(tree.xpath("normalize-space(//table[contains(@data-step, '2')]//td/text())"))
    StaffExperience = int(tree.xpath("normalize-space(//table[contains(@data-step, '4')]//td/text())"))
    StaffMotivation = tree.xpath("//table[contains(@data-step, '4')]/tr[2]/td[1]/text()")[0]
    StaffTechSkill = tree.xpath("//table[contains(@data-step, '4')]/tr[3]/td[1]/text()")[0]
    StaffStressHandling = tree.xpath("//table[contains(@data-step, '4')]/tr[5]/td[1]/text()")[0]
    StaffConcentration = tree.xpath("//table[contains(@data-step, '4')]/tr[6]/td[1]/text()")[0]
    StaffEffiency = tree.xpath("//table[contains(@data-step, '4')]/tr[7]/td[1]/text()")[0]

    #Facilities stuff

    FacWindtunnel = int(tree.xpath("normalize-space(//table[contains(@data-step, '6')]//td/text())"))
    FacPitStop = tree.xpath("//table[contains(@data-step, '6')]/tr[2]/td[1]/text()")[0]
    FacRDWorkshop = tree.xpath("//table[contains(@data-step, '6')]/tr[3]/td[1]/text()")[0]
    FacRDesignCenter = tree.xpath("//table[contains(@data-step, '6')]/tr[4]/td[1]/text()")[0]
    FacEngineerWorkshop = tree.xpath("//table[contains(@data-step, '6')]/tr[5]/td[1]/text()")[0]
    FacAlloyAndChemical = tree.xpath("//table[contains(@data-step, '6')]/tr[6]/td[1]/text()")[0]
    FacCommercial = tree.xpath("//table[contains(@data-step, '6')]/tr[7]/td[1]/text()")[0]

    # Load the workbook
    workbook = openpyxl.load_workbook('me_s80.xlsx')

    # Select the worksheet
    worksheet = workbook['2 Stuff Data']

    current_gp = "Sakhir"

    gp_index = gp_list.index(current_gp)
    col_num = gp_index + 1  # add 2 to account for the header row

    col_num = col_num + 1

    print(col_num)
    # Write the data to the appropriate cell
    worksheet.cell(row=40, column=col_num).value = StaffExperience
    worksheet.cell(row=41, column=col_num).value = StaffMotivation
    worksheet.cell(row=42, column=col_num).value = StaffTechSkill
    worksheet.cell(row=43, column=col_num).value = StaffStressHandling
    worksheet.cell(row=44, column=col_num).value = StaffConcentration
    worksheet.cell(row=45, column=col_num).value = StaffEffiency

    worksheet.cell(row=48, column=col_num).value = FacWindtunnel
    worksheet.cell(row=49, column=col_num).value = FacPitStop
    worksheet.cell(row=50, column=col_num).value = FacRDWorkshop
    worksheet.cell(row=51, column=col_num).value = FacRDesignCenter
    worksheet.cell(row=52, column=col_num).value = FacEngineerWorkshop
    worksheet.cell(row=53, column=col_num).value = FacAlloyAndChemical
    worksheet.cell(row=54, column=col_num).value = FacCommercial

    # Save the changes to the workbook
    workbook.save('me_s80.xlsx')

    response = br.open('https://www.gpro.net/gb/UpdateCar.asp')
    tree = html.fromstring(br.response().get_data())

    ChassisLvl = int(tree.xpath("normalize-space(//td[contains(@id, 'newLvlCha')]/text())"))
    ChassisWear = int(tree.xpath("normalize-space(//td[contains(@id, 'newWearCha')]/text())").strip('%'))
    EngineLvl = int(tree.xpath("normalize-space(//td[contains(@id, 'newLvlEng')]/text())"))
    EngineWear = int(tree.xpath("normalize-space(//td[contains(@id, 'newWearEng')]/text())").strip('%'))
    FWLvl = int(tree.xpath("normalize-space(//td[contains(@id, 'newLvlFW')]/text())"))
    FWWear = int(tree.xpath("normalize-space(//td[contains(@id, 'newWearFW')]/text())").strip('%'))
    RWLvl = int(tree.xpath("normalize-space(//td[contains(@id, 'newLvlRW')]/text())"))
    RWWear = int(tree.xpath("normalize-space(//td[contains(@id, 'newWearRW')]/text())").strip('%'))
    UnderBodyLvl = int(tree.xpath("normalize-space(//td[contains(@id, 'newLvlUB')]/text())"))
    UnderBodyWear = int(tree.xpath("normalize-space(//td[contains(@id, 'newWearUB')]/text())").strip('%'))
    SidePodLvl = int(tree.xpath("normalize-space(//td[contains(@id, 'newLvlSid')]/text())"))
    SidePodWear = int(tree.xpath("normalize-space(//td[contains(@id, 'newWearSid')]/text())").strip('%'))
    CoolingLvl = int(tree.xpath("normalize-space(//td[contains(@id, 'newLvlCoo')]/text())"))
    CoolingWear = int(tree.xpath("normalize-space(//td[contains(@id, 'newWearCoo')]/text())").strip('%'))
    GearBoxLvl = int(tree.xpath("normalize-space(//td[contains(@id, 'newLvlGea')]/text())"))
    GearBoxWear = int(tree.xpath("normalize-space(//td[contains(@id, 'newWearGea')]/text())").strip('%'))
    BrakesLvl = int(tree.xpath("normalize-space(//td[contains(@id, 'newLvlBra')]/text())"))
    BrakesWear = int(tree.xpath("normalize-space(//td[contains(@id, 'newWearBra')]/text())").strip('%'))
    SuspensionLvl = int(tree.xpath("normalize-space(//td[contains(@id, 'newLvlSus')]/text())"))
    SuspensionWear = int(tree.xpath("normalize-space(//td[contains(@id, 'newWearSus')]/text())").strip('%'))
    ElectronicsLvl = int(tree.xpath("normalize-space(//td[contains(@id, 'newLvlEle')]/text())"))
    ElectronicsWear = int(tree.xpath("normalize-space(//td[contains(@id, 'newWearEle')]/text())").strip('%'))

    # Load the workbook
    workbook = openpyxl.load_workbook('me_s80.xlsx')

    # Select the worksheet
    worksheet = workbook['3 Part Data']

    current_gp = "Sakhir"

    gp_index = gp_list.index(current_gp)
    col_num = gp_index + 2  # add 2 to account for the header row

    print(ChassisLvl)

    worksheet.cell(row=5, column=col_num).value = ChassisLvl
    worksheet.cell(row=6, column=col_num).value = EngineLvl
    worksheet.cell(row=7, column=col_num).value = FWLvl
    worksheet.cell(row=8, column=col_num).value = RWLvl
    worksheet.cell(row=9, column=col_num).value = UnderBodyLvl
    worksheet.cell(row=10, column=col_num).value = SidePodLvl
    worksheet.cell(row=11, column=col_num).value = CoolingLvl
    worksheet.cell(row=12, column=col_num).value = GearBoxLvl
    worksheet.cell(row=13, column=col_num).value = BrakesLvl
    worksheet.cell(row=14, column=col_num).value = SuspensionLvl
    worksheet.cell(row=15, column=col_num).value = ElectronicsLvl


    worksheet.cell(row=18, column=col_num).value = ChassisWear / 100
    worksheet.cell(row=19, column=col_num).value = EngineWear / 100
    worksheet.cell(row=20, column=col_num).value = FWWear / 100
    worksheet.cell(row=21, column=col_num).value = RWWear / 100
    worksheet.cell(row=22, column=col_num).value = UnderBodyWear / 100
    worksheet.cell(row=23, column=col_num).value = SidePodWear / 100
    worksheet.cell(row=24, column=col_num).value = CoolingWear / 100
    worksheet.cell(row=25, column=col_num).value = GearBoxWear / 100
    worksheet.cell(row=26, column=col_num).value = BrakesWear / 100
    worksheet.cell(row=27, column=col_num).value = SuspensionWear / 100
    worksheet.cell(row=28, column=col_num).value = ElectronicsWear / 100

    worksheet.cell(row=31, column=col_num).value = ChassisWear / 100
    worksheet.cell(row=32, column=col_num).value = EngineWear / 100
    worksheet.cell(row=33, column=col_num).value = FWWear / 100
    worksheet.cell(row=34, column=col_num).value = RWWear / 100
    worksheet.cell(row=35, column=col_num).value = UnderBodyWear / 100
    worksheet.cell(row=36, column=col_num).value = SidePodWear / 100
    worksheet.cell(row=37, column=col_num).value = CoolingWear / 100
    worksheet.cell(row=38, column=col_num).value = GearBoxWear / 100
    worksheet.cell(row=39, column=col_num).value = BrakesWear / 100
    worksheet.cell(row=40, column=col_num).value = SuspensionWear / 100
    worksheet.cell(row=41, column=col_num).value = ElectronicsWear / 100

    col_num = col_num + 1

    if worksheet.cell(row=44, column=col_num).value is None:
        
        worksheet.cell(row=44, column=col_num).value = ChassisWear / 100
        worksheet.cell(row=45, column=col_num).value = EngineWear / 100
        worksheet.cell(row=46, column=col_num).value = FWWear / 100
        worksheet.cell(row=47, column=col_num).value = RWWear / 100
        worksheet.cell(row=48, column=col_num).value = UnderBodyWear / 100
        worksheet.cell(row=49, column=col_num).value = SidePodWear / 100
        worksheet.cell(row=50, column=col_num).value = CoolingWear / 100
        worksheet.cell(row=51, column=col_num).value = GearBoxWear / 100
        worksheet.cell(row=52, column=col_num).value = BrakesWear / 100
        worksheet.cell(row=53, column=col_num).value = SuspensionWear / 100
        worksheet.cell(row=54, column=col_num).value = ElectronicsWear / 100
        
        col_num = col_num + 1
        worksheet.cell(row=57, column=col_num).value = ChassisWear / 100
        worksheet.cell(row=58, column=col_num).value = EngineWear / 100
        worksheet.cell(row=59, column=col_num).value = FWWear / 100
        worksheet.cell(row=60, column=col_num).value = RWWear / 100
        worksheet.cell(row=61, column=col_num).value = UnderBodyWear / 100
        worksheet.cell(row=62, column=col_num).value = SidePodWear / 100
        worksheet.cell(row=63, column=col_num).value = CoolingWear / 100
        worksheet.cell(row=64, column=col_num).value = GearBoxWear / 100
        worksheet.cell(row=65, column=col_num).value = BrakesWear / 100
        worksheet.cell(row=66, column=col_num).value = SuspensionWear / 100
        worksheet.cell(row=67, column=col_num).value = ElectronicsWear / 100

    col_num = col_num - 1

    if worksheet.cell(row=57, column=col_num).value is None:
        worksheet.cell(row=57, column=col_num).value = ChassisWear / 100
        worksheet.cell(row=58, column=col_num).value = EngineWear / 100
        worksheet.cell(row=59, column=col_num).value = FWWear / 100
        worksheet.cell(row=60, column=col_num).value = RWWear / 100
        worksheet.cell(row=61, column=col_num).value = UnderBodyWear / 100
        worksheet.cell(row=62, column=col_num).value = SidePodWear / 100
        worksheet.cell(row=63, column=col_num).value = CoolingWear / 100
        worksheet.cell(row=64, column=col_num).value = GearBoxWear / 100
        worksheet.cell(row=65, column=col_num).value = BrakesWear / 100
        worksheet.cell(row=66, column=col_num).value = SuspensionWear / 100
        worksheet.cell(row=67, column=col_num).value = ElectronicsWear / 100


    # Save the changes to the workbook
    workbook.save('me_s80.xlsx')

    ## Needs refactor this part of the Part Data
    
if __name__ == '__main__':
    main()

